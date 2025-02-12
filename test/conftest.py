"""Collection of fixtures to assist the testing scripts."""
import configparser
import os
from collections import OrderedDict

import appdirs
import certifi
import openpyxl
import pytest
from dotmap import DotMap

from safedata_validator.field import Dataset, DataWorksheet
from safedata_validator.locations import Locations
from safedata_validator.resources import Resources
from safedata_validator.taxa import (
    GBIFTaxa,
    LocalGBIFValidator,
    LocalNCBIValidator,
    RemoteGBIFValidator,
    RemoteNCBIValidator,
)


def fixture_files():
    """Function to direct test functions to the fixture files.

    Whenever testing is run, this conftest file is loaded and the path of the
    file can be used to provide paths to the location of all other testing and
    fixture files on a given test system. This helps make sure that absolute
    paths are maintained so that testing is not sensitive to the directory where
    tests are run.

    The Dotmap contains real files (.rf) that will be copied into the fake
    filesystem, along with key paths to virtual files (.vf) created within the
    same system. A missing file (.mf) is also provided to test responses to
    missing files.
    """

    fixture_dir = os.path.join(os.path.dirname(__file__), "fixtures")

    real_files = [
        ("loc_file", "locations.json"),
        ("gbif_file", "gbif_backbone_truncated.sqlite"),
        ("ncbi_file", "ncbi_database_truncated.sqlite"),
        ("sqlite_not_gbif", "database_file_not_gbif.sqlite"),
        ("json_not_locations", "notalocationsjson.json"),
        ("bad_excel_file", "Test_format_bad.xlsx"),
        ("good_excel_file", "Test_format_good.xlsx"),
        ("bad_ncbi_file", "Test_format_bad_NCBI.xlsx"),
        ("weird_ncbi_file", "Test_format_weird_NCBI.xlsx"),
        ("good_ncbi_file", "Test_format_good_NCBI.xlsx"),
    ]

    # Check if testing is part of continuos integration
    on_github_action = os.getenv("CI", False)

    if not on_github_action:
        real_files.append(("config_secrets", "config_secrets.ini"))

    real_files = {ky: os.path.join(fixture_dir, vl) for ky, vl in real_files}

    # Check that a secrets.ini file has actually been provided
    if not on_github_action:
        if not os.path.exists(real_files["config_secrets"]):
            raise OSError(
                "config_secrets.ini not found. Use fixtures/config_secrets_template.ini"
                " to create a config_secrets.ini file for configuring local pytests"
            )

    # Need to provide the path to the certifi CA bundle or requests breaks!
    real_files["certifi"] = certifi.where()

    virtual_files = {
        "user_config": os.path.join(
            appdirs.user_config_dir(), "safedata_validator", "safedata_validator.cfg"
        ),
        "site_config": os.path.join(
            appdirs.site_config_dir(), "safedata_validator", "safedata_validator.cfg"
        ),
        "fix_cfg_local": os.path.join(fixture_dir, "safedata_validator_local.cfg"),
        "fix_cfg_remote": os.path.join(fixture_dir, "safedata_validator_remote.cfg"),
        "fix_cfg_local_ncbi": os.path.join(
            fixture_dir, "safedata_validator_local_ncbi.cfg"
        ),
    }

    return DotMap(
        dict(
            rf=real_files,
            vf=virtual_files,
            mf=os.path.join(fixture_dir, "thisfiledoesnotexist"),
        )
    )


FIXTURE_FILES = fixture_files()
"""DotMap: This global variable contains a dotmap of the paths of various
key testing files. This could be a fixture, but these paths are used with
parameterisation as well as within tests, and using fixture values in
parameterisation is clumsy and complex.
"""


def log_check(caplog, expected_log):
    """Helper function to check that the captured log is as expected.

    Arguments:
        caplog: An instance of the caplog fixture
        expected_log: An iterable of 2-tuples containing the
            log level and message.
    """

    assert len(expected_log) == len(caplog.records)

    assert all(
        [exp[0] == rec.levelno for exp, rec in zip(expected_log, caplog.records)]
    )
    assert all(
        [exp[1] in rec.message for exp, rec in zip(expected_log, caplog.records)]
    )


"""
The other contents are fixtures that will be available to all test suites.


"""


@pytest.fixture()
def config_filesystem(fs):
    """Create a config and resources file system for testing.

    Testing requires access to the configuration files for the package resources
    and the paths to these files are going to differ across test system. This
    fixture uses the pyfakefs plugin for pytest to create a fake file system
    containing config files and versions of the resources, cut down to save
    filespace.

    See also: https://gist.github.com/peterhurford/09f7dcda0ab04b95c026c60fa49c2a68

    Args:
        fs: The pyfakefs plugin object

    Returns:
        A fake filesystem containing copies of key actual fixture files in their
        correct location, along with virtual config files containing the correct
        paths to those fixtures files.
    """

    # Point to real locations of test fixture example resources
    for _, val in FIXTURE_FILES.rf.items():
        fs.add_real_file(val)

    # Check if tests are being run as part of continuos integration
    on_github_action = os.getenv("CI", False)
    ncbi_api_key = None

    if on_github_action:
        ncbi_api_key = os.getenv("NCBI_API_KEY")
    else:
        conf = configparser.ConfigParser()
        conf.read(FIXTURE_FILES.rf.config_secrets)
        ncbi_api_key = conf["ncbi"]["api_key"]

    if ncbi_api_key is None:
        raise RuntimeError("NCBI_API_KEY not found")

    # The config files _cannot_ be real files because they need to
    # contain absolute paths to the resources available on the specific
    # test machine, and so need to be created with those paths

    config_contents = [
        "locations = ",
        "gbif_database = ",
        "ncbi_database = ",
        "[extents]",
        "temporal_soft_extent = 2002-02-02, 2030-01-31",
        "temporal_hard_extent = 2002-02-01, 2030-02-01",
        "latitudinal_hard_extent = -90, 90",
        "latitudinal_soft_extent = -4, 8",
        "longitudinal_hard_extent = -180, 180",
        "longitudinal_soft_extent = 110, 120",
        "[zenodo]",
        "community_name = safe",
        "use_sandbox = True",
        "zenodo_sandbox_api = https://sandbox.zenodo.org",
        "zenodo_sandbox_token = xyz",
        "zenodo_api = https://api.zenodo.org",
        "zenodo_token = xyz",
        "[ncbi]",
        f"api_key = {ncbi_api_key}",
    ]

    # Remote config (no gbif database) in the fixture directory
    config_contents[0] += FIXTURE_FILES.rf.loc_file
    fs.create_file(FIXTURE_FILES.vf.fix_cfg_remote, contents="\n".join(config_contents))

    # Local config (local GBIF database) in the fixture directory
    config_contents[1] += FIXTURE_FILES.rf.gbif_file
    fs.create_file(FIXTURE_FILES.vf.fix_cfg_local, contents="\n".join(config_contents))

    # Local config (local NCBI database) in the fixture directory
    config_contents[2] += FIXTURE_FILES.rf.ncbi_file
    fs.create_file(
        FIXTURE_FILES.vf.fix_cfg_local_ncbi, contents="\n".join(config_contents)
    )

    # # Local user config
    # fs.create_file(FIXTURE_FILES.vf.user_config, contents='\n'.join(config_contents))

    yield fs


@pytest.fixture()
def user_config_file(config_filesystem):
    """Attempt to extend the fake fs - does not seem to work."""
    # Local user config as a duplicate of the existing config in the fixture
    # directory

    with open(FIXTURE_FILES.vf.fix_cfg_local_ncbi) as infile:
        config_filesystem.create_file(
            FIXTURE_FILES.vf.user_config, contents="".join(infile.readlines())
        )

    yield config_filesystem


@pytest.fixture()
def site_config_file(config_filesystem):
    """Attempt to extend the fake fs - does not seem to work."""
    # Local user config as a duplicate of the existing config in the fixture
    # directory

    with open(FIXTURE_FILES.vf.fix_cfg_local_ncbi) as infile:
        config_filesystem.create_file(
            FIXTURE_FILES.vf.site_config, contents="".join(infile.readlines())
        )

    yield config_filesystem


# ------------------------------------------
# Resources: fixtures containing instances of both local and remote GBIF
#            validators and a parameterised fixture providing both for
#            duplicating tests on both systems.
# ------------------------------------------


@pytest.fixture()
def resources_with_local_gbif(config_filesystem):
    """Creates a Resource object configured to use a local GBIF database.

    Returns:
        A safedata_validator.resources.Resources instance
    """

    return Resources(config=FIXTURE_FILES.vf.fix_cfg_local)


@pytest.fixture()
def resources_with_remote_gbif(config_filesystem):
    """Creates a Resource object configured to use the remote GBIF API.

    Returns:
        A safedata_validator.resources.Resources instance
    """

    return Resources(config=FIXTURE_FILES.vf.fix_cfg_remote)


@pytest.fixture(params=["remote", "local"])
def resources_local_and_remote(
    request, resources_with_local_gbif, resources_with_remote_gbif
):
    """Parameterised fixture to run tests using both the local and remote GBIF."""

    if request.param == "remote":
        if os.getenv("SDV_NO_REMOTE") in ["1", "true", "True"]:
            pytest.skip("Remote testing turned off via SDV_NO_REMOTE")
        else:
            return resources_with_remote_gbif
    elif request.param == "local":
        return resources_with_local_gbif


@pytest.fixture()
def resources_with_local_ncbi(config_filesystem):
    """Creates a Resource object configured to use a local NCBI database.

    Returns:
        A safedata_validator.resources.Resources instance
    """

    return Resources(config=FIXTURE_FILES.vf.fix_cfg_local_ncbi)


@pytest.fixture()
def resources_with_remote_ncbi(config_filesystem):
    """Creates a Resource object configured to use the remote NCBI API.

    Returns:
        A safedata_validator.resources.Resources instance
    """

    return Resources(config=FIXTURE_FILES.vf.fix_cfg_remote)


@pytest.fixture(params=["remote", "local"])
def ncbi_resources_local_and_remote(
    request, resources_with_local_ncbi, resources_with_remote_ncbi
):
    """Parameterised fixture to run tests using both the local and remote NCBI."""

    if request.param == "remote":
        if os.getenv("SDV_NO_REMOTE") in ["1", "true", "True"]:
            pytest.skip("Remote testing turned off via SDV_NO_REMOTE")
        else:
            return resources_with_remote_ncbi
    elif request.param == "local":
        return resources_with_local_ncbi


# ------------------------------------------
# Other fixtures
# ------------------------------------------


@pytest.fixture()
def example_excel_files(config_filesystem, request):
    """Fixture providing example excel files for testing.

    This uses indirect parameterisation, to allow the shared fixture to be paired with
    request specific expectations rather than all pair combinations:

    https://stackoverflow.com/questions/70379640
    """
    if request.param == "good":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.good_excel_file, read_only=True)
        return wb
    elif request.param == "bad":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.bad_excel_file, read_only=True)
        return wb


@pytest.fixture()
def example_ncbi_files(config_filesystem, request):
    """Fixture providing additional excel files with NCBI taxa details for testing.

    This uses indirect parameterisation, to allow the shared fixture to be paired with
    request specific expectations rather than all pair combinations. This is an
    equivalent function to example_excel_files but for the NCBI specific excel files.
    """
    if request.param == "good":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.good_ncbi_file, read_only=True)
        return wb
    elif request.param == "bad":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.bad_ncbi_file, read_only=True)
        return wb
    elif request.param == "weird":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.weird_ncbi_file, read_only=True)
        return wb


@pytest.fixture(params=["remote", "local"])
def fixture_taxon_validators(resources_with_local_gbif, request):
    """Parameterised fixture to return local and remote taxon validators."""
    if request.param == "remote":
        if os.getenv("SDV_NO_REMOTE") in ["1", "true", "True"]:
            pytest.skip("Remote testing turned off via SDV_NO_REMOTE")
        else:
            return RemoteGBIFValidator()

    elif request.param == "local":
        return LocalGBIFValidator(resources_with_local_gbif)


@pytest.fixture(params=["remote", "local"])
def fixture_ncbi_validators(
    resources_with_local_ncbi, resources_with_remote_ncbi, request
):
    """Parameterised fixture to return local and remote NCBI validator."""
    if request.param == "remote":
        if os.getenv("SDV_NO_REMOTE") in ["1", "true", "True"]:
            pytest.skip("Remote testing turned off via SDV_NO_REMOTE")
        else:
            return RemoteNCBIValidator(resources_with_remote_ncbi)

    elif request.param == "local":
        return LocalNCBIValidator(resources_with_local_ncbi)


# Fixtures to provide Taxon, Locations, Dataset, Dataworksheet
# and field meta objects for testing


@pytest.fixture()
def fixture_taxa(resources_with_local_gbif):
    """Fixture to provide a taxon object with a couple of names.

    These examples need to be in the cutdown local GBIF testing database in fixtures.
    """

    taxa = GBIFTaxa(resources_with_local_gbif)

    test_taxa = [
        ("C_born", ["Crematogaster borneensis", "Species", None, None], None),
        ("V_salv", ["Varanus salvator", "Species", None, None], None),
    ]

    for tx in test_taxa:
        taxa.validate_and_add_taxon(tx)

    return taxa


@pytest.fixture()
def fixture_locations(resources_with_local_gbif):
    """Fixture to provide a taxon object with a couple of names.

    These examples need to be in the cutdown local GBIF testing database in fixtures.
    """

    locations = Locations(resources_with_local_gbif)

    test_locs = ["A_1", "A_2", 1, 2]

    locations.add_known_locations(test_locs)

    return locations


@pytest.fixture()
def fixture_dataset(resources_with_local_gbif):
    """Fixture to provide a dataset.

    This dataset has been prepopulated with some taxon and location names for field
    tests.
    """

    dataset = Dataset(resources_with_local_gbif)

    test_taxa = [
        ("C_born", ["Crematogaster borneensis", "Species", None, None], None),
        ("V_salv", ["Varanus salvator", "Species", None, None], None),
    ]

    for tx in test_taxa:
        dataset.taxa.gbif_taxa.validate_and_add_taxon(tx)

    test_locs = ["A_1", "A_2", 1, 2]

    dataset.locations.add_known_locations(test_locs)

    return dataset


@pytest.fixture(scope="module")
def fixture_field_meta():
    """field_meta object for use across tests."""

    return OrderedDict(
        field_type=["numeric", "numeric", "numeric"],
        description=["a", "b", "c"],
        units=["a", "b", "c"],
        method=["a", "b", "c"],
        field_name=["a", "b", "c"],
    )


@pytest.fixture()
def fixture_dataworksheet(fixture_dataset):
    """field_meta object for use across tests."""

    dws = DataWorksheet(
        {
            "name": "DF",
            "title": "My data table",
            "description": "This is a test data worksheet",
            "external": None,
        },
        dataset=fixture_dataset,
    )

    return dws
